"""
Core logic — อ่าน Excel ตารางเรือ + วาดปฏิทิน PNG
(ย้ายมาจาก Ship_Schedule.py เวอร์ชัน tkinter — logic เดิมทุกอย่าง
 ต่างแค่รูปเรือ/โลโก้อ่านจาก static/img แทน base64 ที่ฝังในไฟล์)
"""
import re
from calendar import month_name, Calendar
from datetime import date, datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFont

IMG_DIR = Path(__file__).parent / "static" / "img"

# ─── PARSE EXCEL ──────────────────────────────────────────────────────────────
MONTHS = {m[:3].lower(): i for i, m in enumerate(month_name) if m}
HEADER_ROW, DATA_START = 2, 3
COL = {
    'vessel':  'Vessel Name',
    'arrival': 'Estimated Arrival Date',
    'berth':   'Berth',
    'lm1':     'Loading Master 1',
    'lm2':     'Loading Master 2',
    'yy':      'YY',
    'mm':      'MM',
    'shipper': 'Shipper',
}

def parse_arrival(val):
    """คืน (day, year) จาก cell — year อาจเป็น None ถ้าดึงไม่ได้"""
    if isinstance(val, datetime): return val.day, val.year
    if isinstance(val, date):     return val.day, val.year
    if isinstance(val, (int, float)):
        try:
            d = datetime(1899, 12, 30) + timedelta(days=float(val))
            return d.day, d.year
        except Exception: return None, None
    if isinstance(val, str):
        s = val.strip()
        m = re.match(r'^(\d{1,2})\s*-\s*\d{1,2}\s+[A-Za-z]{3}[\s\-]+(\d{2,4})', s)
        if m:
            ys = m.group(2)
            return int(m.group(1)), int(ys)+2000 if len(ys) == 2 else int(ys)
        m = re.match(r'^(\d{1,2})[\-/][A-Za-z]{3}[\-\s]+(\d{2,4})', s)
        if m:
            ys = m.group(2)
            return int(m.group(1)), int(ys)+2000 if len(ys) == 2 else int(ys)
        m = re.match(r'^(\d{1,2})', s)
        if m: return int(m.group(1)), None
    return None, None

def parse_lm(raw):
    s = raw.strip()
    m = re.search(r'\(([^)]+)\)', s)
    return m.group(1).strip() if m else s

def resolve_logo_key(shipper):
    # เช็ค PTT ท้ายสุด เพื่อให้ shipper ผสมอย่าง PTT&PGL ได้โลโก้ PGL ก่อน
    s = shipper.upper()
    for key in ('HKH', 'GLNG', 'BGM', 'PGL', 'EGT', 'PTT'):
        if key in s: return key.lower()
    return None

def _first_number(row, cols):
    """ค่าตัวเลขแรกที่ไม่ใช่ 0 จากรายการ index คอลัมน์ (ไม่มีคืน None)"""
    for i in cols:
        try:
            f = float(row[i])
            if f: return f
        except (TypeError, ValueError):
            pass
    return None

def _marked(val):
    return bool(str(val or '').strip())

def available_years(excel_path, sheet):
    """คืนรายการปีทั้งหมดที่มีข้อมูลในไฟล์ Excel (เรียงน้อยไปมาก)"""
    wb = load_workbook(excel_path, data_only=True)
    if sheet not in wb.sheetnames:
        raise ValueError(f"ไม่พบ sheet '{sheet}' — มี: {wb.sheetnames}")
    ws = wb[sheet]
    hdrs = [(c.value or '').replace('\n', ' ').strip() for c in ws[HEADER_ROW]]
    low = [h.lower() for h in hdrs]
    try:
        ci_yy = low.index('yy')
        ci_vessel = low.index(COL['vessel'].lower())
    except ValueError:
        raise ValueError("ไม่พบคอลัมน์ 'YY' หรือ 'Vessel Name'")
    years = set()
    for row in ws.iter_rows(min_row=DATA_START, values_only=True):
        if not row[ci_vessel]:
            continue
        try:
            y = int(row[ci_yy])
        except (TypeError, ValueError):
            continue
        if 2000 <= y <= 2100:
            years.add(y)
    return sorted(years)

def load_vessels(excel_path, sheet, year, months):
    """อ่านเรือของปี `year` เฉพาะเดือนใน `months` (iterable ของเลขเดือน 1-12)
    คืน (vessels, skipped) — แต่ละลำมี key 'month', skipped เป็น (month, ชื่อเรือ)"""
    months = set(months)
    wb = load_workbook(excel_path, data_only=True)
    if sheet not in wb.sheetnames:
        raise ValueError(f"ไม่พบ sheet '{sheet}' — มี: {wb.sheetnames}")
    ws = wb[sheet]
    hdrs = [(c.value or '').replace('\n', ' ').strip() for c in ws[HEADER_ROW]]
    ci = {}
    for k, label in COL.items():
        for i, h in enumerate(hdrs):
            if h.lower() == label.lower():
                ci[k] = i; break
        if k not in ci:
            raise ValueError(f"ไม่พบ column '{label}'")
    # คอลัมน์เสริม (ไม่มีในไฟล์ก็ไม่ error) — ใน Excel มี Unloading Quantity /
    # Density อย่างละ 2 คอลัมน์ จึงเก็บมาทุกคอลัมน์แล้วใช้ค่าแรกที่ไม่ใช่ 0
    low = [h.lower() for h in hdrs]
    qty_cols  = [i for i, h in enumerate(low) if h.startswith('unloading q')]
    dens_cols = [i for i, h in enumerate(low) if h == 'density']
    ci_nom  = next((i for i, h in enumerate(low) if h == 'nom type'), None)
    ci_lt   = next((i for i, h in enumerate(low) if h == 'lt'), None)
    ci_spot = next((i for i, h in enumerate(low) if h == 'spot'), None)
    vessels, skipped = [], []
    for row in ws.iter_rows(min_row=DATA_START, values_only=True):
        if not row[ci['vessel']]: continue
        yy, mm = row[ci['yy']], row[ci['mm']]
        if yy is None or mm is None: continue
        try: yy = int(yy)
        except Exception: continue
        mm_n = MONTHS.get(str(mm).strip().lower()[:3])
        if yy != year or mm_n not in months: continue
        day, arr_year = parse_arrival(row[ci['arrival']])
        if day is None:
            skipped.append((mm_n, str(row[ci['vessel']]))); continue
        # ข้าม preliminary: arrival ไม่มีปี
        if arr_year is None:
            continue
        # ข้าม preliminary: ปีไม่ตรงกับที่เลือก
        if arr_year != year:
            skipped.append((mm_n, str(row[ci['vessel']]))); continue
        # ข้าม preliminary: ยังไม่มี shipper (เรือที่ confirm แล้วต้องมี shipper เสมอ)
        raw_shipper = str(row[ci['shipper']] or '').strip()
        if not raw_shipper or raw_shipper.lower() == 'none':
            continue
        try: berth = int(row[ci['berth']])
        except Exception: berth = 1
        # ประเภทสัญญา: ใช้คอลัมน์ Nom type ก่อน ถ้าว่างดูเครื่องหมายในช่อง LT/Spot
        nom = str(row[ci_nom] or '').strip() if ci_nom is not None else ''
        if not nom:
            if ci_lt is not None and _marked(row[ci_lt]):     nom = 'LT'
            elif ci_spot is not None and _marked(row[ci_spot]): nom = 'Spot'
        vessels.append({
            'month':   mm_n,
            'day':     day,
            'vessel':  str(row[ci['vessel']]).strip(),
            'berth':   berth,
            'lm1':     parse_lm(str(row[ci['lm1']] or '')),
            'lm2':     parse_lm(str(row[ci['lm2']] or '')),
            'shipper': raw_shipper.upper(),
            'nom':     nom,
            'qty':     _first_number(row, qty_cols),
            'density': _first_number(row, dens_cols),
        })
    return sorted(vessels, key=lambda v: (v['month'], v['day'])), skipped


# ─── DRAW PNG ─────────────────────────────────────────────────────────────────
_CACHE: dict = {}

def _get_img(key):
    if key in _CACHE: return _CACHE[key]
    p = IMG_DIR / f"{key}.png"
    if not p.exists():
        _CACHE[key] = None; return None
    img = Image.open(p).convert("RGBA")
    _CACHE[key] = img; return img

def get_font(size, bold=False):
    for p in [
        "C:/Windows/Fonts/arialbd.ttf" if bold else "C:/Windows/Fonts/arial.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf" if bold else
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
    ]:
        try: return ImageFont.truetype(p, size)
        except Exception: pass
    return ImageFont.load_default()

def paste_ship(canvas, cx, top_y, target_w, berth):
    src = _get_img(f'ship_berth{berth}')
    if src is None: return 36
    ow, oh = src.size
    nh = int(oh * target_w / ow)
    ship = src.resize((target_w, nh), Image.LANCZOS)
    canvas.paste(ship, (int(cx - target_w/2), int(top_y)), mask=ship.split()[3])
    return nh

def paste_logo(canvas, cx, ship_top, ship_h, logo_key):
    src = _get_img(f'logo_{logo_key}')
    if src is None: return
    mh = max(int(ship_h * 0.55), 12)
    mw = max(int(ship_h * 1.6),  30)
    logo = src.copy()
    logo.thumbnail((mw, mh), Image.LANCZOS)
    lw, lh = logo.size
    canvas.paste(logo, (int(cx - lw/2), int(ship_top + ship_h * 0.25)),
                 mask=logo.split()[3])

def dtc(draw, cx, y, text, font, fill):
    bb = draw.textbbox((0, 0), text, font=font)
    draw.text((cx - (bb[2]-bb[0])/2, y), text, font=font, fill=fill)

def build_calendar(vessels, month, year, update_str, W=1500, H=950):
    img  = Image.new('RGB', (W, H), 'white')
    draw = ImageDraw.Draw(img)
    RED  = (200, 30, 30); LGRAY = (240, 240, 240)
    MGRAY= (180,180,180); BLK   = (20,  20,  20); NAVY = (20, 35, 100)
    mx   = 36
    f_title = get_font(52, True); f_year  = get_font(38, True)
    f_sub   = get_font(22);        f_wd    = get_font(18, True)
    f_day   = get_font(17, True);  f_code  = get_font(12.5)
    f_leg   = get_font(15, True);  f_upd   = get_font(19, True)
    f_updv  = get_font(18, True)

    draw.text((mx, 22), month_name[month].upper(), font=f_title, fill=RED)
    bb = draw.textbbox((0,0), str(year), font=f_year)
    draw.text((W-mx-(bb[2]-bb[0]), 22), str(year), font=f_year, fill=RED)
    bb = draw.textbbox((0,0), 'CALENDAR', font=f_sub)
    draw.text((W-mx-(bb[2]-bb[0]), 68), 'CALENDAR', font=f_sub, fill=(50,50,50))
    draw.line([(mx, 108),(W-mx, 108)], fill=MGRAY, width=1)

    WD = ['SUN','MON','TUE','WED','THU','FRI','SAT']
    gt, gb, gl, gr = 118, H-120, mx, W-mx
    cw = (gr-gl)/7; hh = 44
    weeks = Calendar(firstweekday=6).monthdayscalendar(year, month)
    rh = (gb-gt-hh) / len(weeks)

    for i, wd in enumerate(WD):
        x0 = gl + i*cw
        draw.rectangle([x0,gt,x0+cw,gt+hh], fill=(90,20,20) if i==0 else (55,55,55))
        dtc(draw, x0+cw/2, gt+hh/2-9, wd, f_wd, 'white')

    vbd = {}
    for v in vessels: vbd.setdefault(v['day'], []).append(v)

    for r, week in enumerate(weeks):
        for c, day in enumerate(week):
            x0=gl+c*cw; y0=gt+hh+r*rh; x1,y1=x0+cw,y0+rh
            if day == 0:
                draw.rectangle([x0,y0,x1,y1], fill=(158,158,158)); continue
            draw.rectangle([x0,y0,x1,y1],
                           fill=LGRAY if c%2==0 else 'white', outline=MGRAY)
            draw.text((x0+8, y0+6), str(day), font=f_day,
                      fill=RED if c==0 else BLK)
            dvs = vbd.get(day, [])
            if not dvs: continue

            cx = (x0+x1)/2

            if len(dvs) == 1:
                # ─── เรือ 1 ลำ ───────────────────────────────────
                v = dvs[0]
                itop = y0 + rh*0.22; sw = int((cw-18)*0.75)
                sh = paste_ship(img, cx, itop, sw, v['berth'])
                lk = resolve_logo_key(v['shipper'])
                if lk: paste_logo(img, cx, itop, sh, lk)
                name = v['vessel']; ny = itop + sh + 4
                for sz in (12.5, 11, 10, 9, 8, 7):
                    fn = get_font(sz, True)
                    bb = draw.textbbox((0,0), name, font=fn)
                    if bb[2]-bb[0] <= cw-8: break
                dtc(draw, cx, ny, name, fn, BLK)
                dtc(draw, cx, ny+16, f"{v['lm1']}/{v['lm2']}", f_code, (80,80,80))

            elif len(dvs) == 2:
                # ─── เรือ 2 ลำ: วางซ้ายขวา ─────────────────────
                sw2 = int((cw/2 - 14) * 0.85)
                fn2 = get_font(8, True)
                fc2 = get_font(8)
                for idx, v in enumerate(dvs):
                    vcx = x0 + cw*0.25 + idx*cw*0.5
                    itop = y0 + rh*0.22
                    sh = paste_ship(img, vcx, itop, sw2, v['berth'])
                    lk = resolve_logo_key(v['shipper'])
                    if lk: paste_logo(img, vcx, itop, sh, lk)
                    name = v['vessel']; ny = itop + sh + 3
                    half_w = cw/2 - 8
                    for sz in (9, 8, 7):
                        fn2 = get_font(sz, True)
                        bb = draw.textbbox((0,0), name, font=fn2)
                        if bb[2]-bb[0] <= half_w: break
                    dtc(draw, vcx, ny, name, fn2, BLK)
                    dtc(draw, vcx, ny+12, f"{v['lm1']}/{v['lm2']}", fc2, (80,80,80))
                mid = x0 + cw/2
                draw.line([(mid, y0+rh*0.15), (mid, y1-4)], fill=(210,210,210), width=1)

            else:
                # ─── เรือ 3+ ลำ: แสดงลำแรก + "+N more" ──────────
                v = dvs[0]
                itop = y0 + rh*0.22; sw = int((cw-18)*0.75)
                sh = paste_ship(img, cx, itop, sw, v['berth'])
                lk = resolve_logo_key(v['shipper'])
                if lk: paste_logo(img, cx, itop, sh, lk)
                name = v['vessel']; ny = itop + sh + 4
                for sz in (12.5, 11, 10, 9, 8, 7):
                    fn = get_font(sz, True)
                    bb = draw.textbbox((0,0), name, font=fn)
                    if bb[2]-bb[0] <= cw-8: break
                dtc(draw, cx, ny, name, fn, BLK)
                dtc(draw, cx, ny+16, f"{v['lm1']}/{v['lm2']}", f_code, (80,80,80))
                dtc(draw, cx, ny+30, f"+{len(dvs)-1} more", f_code, RED)

    draw.rectangle([gl,gt,gr,gb], outline=MGRAY, width=1)
    lx = W-mx-260; ly = H-108
    paste_ship(img, lx+55, ly,    110, 1)
    draw.text((lx+120, ly+10), 'Berth 1', font=f_leg, fill=NAVY)
    paste_ship(img, lx+55, ly+48, 110, 2)
    draw.text((lx+120, ly+58), 'Berth 2', font=f_leg, fill=NAVY)
    draw.text((mx, H-68), 'Update',     font=f_upd,  fill=RED)
    draw.text((mx, H-42), update_str,   font=f_updv, fill=NAVY)
    return img
