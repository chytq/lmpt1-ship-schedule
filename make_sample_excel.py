"""
สร้างไฟล์ Excel ตัวอย่าง (ข้อมูลสมมติ) สำหรับ dev ที่ยังไม่มีไฟล์จริง

    python make_sample_excel.py

จะได้ไฟล์ sample_work_plan.xlsx ที่มีโครงสร้างคอลัมน์เหมือนไฟล์จริงทุกอย่าง
พอมีไฟล์นี้แล้ว รัน `python app.py` ได้เลยโดยไม่ต้องใช้ข้อมูลจริง
"""
import random
from calendar import monthrange
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook

OUT = Path(__file__).parent / "sample_work_plan.xlsx"

# หัวตารางต้องตรงกับ COL ใน core.py (แถวที่ 2, ข้อมูลเริ่มแถวที่ 3)
HEADERS = [
    "T1 Terminal Cargo No.", "Customer  Cargo No.", "Shipper", "Cargo size (MMBTU)",
    "Vessel Name", "Estimated Arrival Date", "Berth", "Loading Master 1",
    "Loading Master 2", "LT", "Spot", "Date", "ETA", "YY", "MM",
    "Est. quantity", "Unloading Qunatity", "Unloading Quantity",
    "Density", "Density", "Unloading temp.", "Unloading temp.", "Nom type",
]

VESSELS = [
    "PACIFIC BREEZE", "NORTHERN STAR", "CORAL VOYAGER", "SIAM PIONEER",
    "ANDAMAN SPIRIT", "GULF HARMONY", "AZURE DAWN", "MEKONG TRADER",
    "EASTERN PEARL", "BLUE HORIZON", "SUMMIT GLORY", "OCEAN SERENADE",
]
SHIPPERS = ["PTT", "PGL", "GLNG", "BGM", "HKH", "EGT"]
MASTERS = ["AB", "CD", "EF", "GH", "IJ", "KL"]
MONTHS3 = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
           "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]


def build_rows(years):
    rng = random.Random(20260810)          # fix seed ให้ผลลัพธ์เหมือนเดิมทุกครั้ง
    rows = []
    cargo = 1
    today = date.today()
    for year in years:
        for m in range(1, 13):
            last_day = monthrange(year, m)[1]
            days = sorted(rng.sample(range(1, last_day + 1), rng.randint(4, 8)))
            for d in days:
                shipper = rng.choice(SHIPPERS)
                nom = rng.choice(["LT", "LT", "LT", "Spot"])
                arrived = date(year, m, d) < today
                qty = rng.randint(125_000, 175_000) if arrived else 0
                dens = round(rng.uniform(415, 460), 2) if arrived else 0
                lm1, lm2 = rng.sample(MASTERS, 2)
                rows.append([
                    f"T1-{cargo:04d}", f"C-{cargo:04d}", shipper,
                    rng.randint(3_000_000, 3_800_000),
                    rng.choice(VESSELS),
                    datetime(year, m, d),                 # Estimated Arrival Date
                    rng.choice([1, 2]),                   # Berth
                    f"Mr. Somchai ({lm1})", f"Mr. Prasert ({lm2})",
                    "x" if nom == "LT" else None,
                    "x" if nom == "Spot" else None,
                    datetime(year, m, d), "08:00",
                    year, MONTHS3[m - 1],
                    rng.randint(3_000_000, 3_800_000),
                    qty, qty, dens, dens,
                    -160 if arrived else 0, -160 if arrived else 0,
                    nom,
                ])
                cargo += 1
    return rows


def main():
    this_year = date.today().year
    years = [this_year - 1, this_year]

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append([])                       # แถว 1 เว้นว่าง (เหมือนไฟล์จริง)
    ws.append(HEADERS)                  # แถว 2 = หัวตาราง
    for row in build_rows(years):       # แถว 3 เป็นต้นไป = ข้อมูล
        ws.append(row)
    for cell in ws[2]:
        cell.style = "Headline 3"
    wb.save(OUT)

    print(f"สร้างแล้ว: {OUT.name}")
    print(f"  ปี      : {', '.join(map(str, years))}")
    print(f"  จำนวนเรือ: {ws.max_row - 2} ลำ")
    print()
    print("รันเว็บด้วย:  python app.py   แล้วเปิด http://localhost:5000")


if __name__ == "__main__":
    main()
